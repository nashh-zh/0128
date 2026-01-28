import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import numpy as np
from datetime import datetime, timedelta
import os
import json
import shutil
import logging
import pickle
import sys
import platform

# 图表相关
import matplotlib

# 修改后端：在 macOS 上 TkAgg 通常比 Agg 嵌入效果更好，且支持交互
try:
    matplotlib.use('TkAgg')
except:
    matplotlib.use('Agg')  # 回退方案

import matplotlib.pyplot as plt
from matplotlib.figure import Figure
from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk

# Excel 样式相关
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import DataBarRule

# 忽略警告
import warnings

warnings.filterwarnings('ignore')


class ProfessionalMarginCalculator:
    def __init__(self, root):
        self.root = root
        self.root.title("📊 专业销售毛利分析系统 v4.4 (Mac版)")

        # macOS 窗口初始大小
        self.root.geometry("1400x900")

        # 检测操作系统以选择最佳字体
        self.system = platform.system()
        if self.system == "Darwin":  # macOS
            self.default_font = "PingFang SC"
            self.font_fallbacks = ['PingFang SC', 'Heiti TC', 'STHeiti', 'Arial Unicode MS', 'Microsoft YaHei',
                                   'SimHei']
        elif self.system == "Windows":
            self.default_font = "Microsoft YaHei"
            self.font_fallbacks = ['Microsoft YaHei', 'SimHei', 'Arial Unicode MS']
        else:
            self.default_font = "DejaVu Sans"
            self.font_fallbacks = ['DejaVu Sans', 'WenQuanYi Micro Hei']

        # 设置日志
        self.setup_logging()

        # 配置设置
        self.config_file = os.path.join(os.path.expanduser("~"), '.margin_analyzer_config.json')
        self.history_data_file = os.path.join(os.path.expanduser("~"), '.margin_history_data.pkl')
        self.load_config()

        # 颜色方案 - 现代化设计
        self.colors = {
            'primary': '#3f51b5',
            'secondary': '#673ab7',
            'accent': '#2196f3',
            'success': '#4caf50',
            'warning': '#ff9800',
            'danger': '#f44336',
            'light': '#f5f5f5',
            'lighter': '#fafafa',
            'dark': '#333333',
            'white': '#ffffff',
            'border': '#e0e0e0'
        }

        self.root.configure(bg=self.colors['lighter'])

        # 数据变量
        self.sales_file_path = tk.StringVar()
        self.latest_purchase_file_path = tk.StringVar()
        self.last_output_path = None
        self.current_detailed_data = None

        # 历史采购数据
        self.history_purchase_data = None
        self.history_loaded = False
        self.history_file_path = tk.StringVar(value="历史采购数据未加载")

        # 累计数据存储
        self.monthly_data = {}
        self.yearly_data = {}
        self.load_cumulative_data()

        # 尝试加载已保存的历史数据
        self.load_history_from_disk()

        # 创建UI组件
        self.setup_menu()
        self.setup_ui()

    def setup_logging(self):
        """设置日志系统"""
        try:
            log_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "logs")
            os.makedirs(log_dir, exist_ok=True)
        except Exception:
            log_dir = os.path.expanduser("~")

        log_file = os.path.join(log_dir, f"sales_analysis_{datetime.now().strftime('%Y%m%d')}.log")

        self.logger = logging.getLogger('SalesAnalysis')
        self.logger.setLevel(logging.INFO)

        if not self.logger.handlers:
            file_handler = logging.FileHandler(log_file, encoding='utf-8')
            file_handler.setLevel(logging.INFO)

            console_handler = logging.StreamHandler()
            console_handler.setLevel(logging.INFO)

            formatter = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            file_handler.setFormatter(formatter)
            console_handler.setFormatter(formatter)

            self.logger.addHandler(file_handler)
            self.logger.addHandler(console_handler)

        self.logger.info("=" * 50)
        self.logger.info(f"销售毛利分析系统启动 (系统: {self.system})")
        self.logger.info("=" * 50)

    def load_config(self):
        """加载配置"""
        default_config = {
            'export_path': os.path.join(os.path.expanduser("~"), "Documents", "销售分析报告"),
            'auto_open': True,
            'create_subfolders': True,
            'date_format': '%Y-%m-%d',
            'remember_history': True
        }

        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    self.config = json.load(f)
            except:
                self.config = default_config
        else:
            self.config = default_config

        try:
            os.makedirs(self.config['export_path'], exist_ok=True)
        except Exception as e:
            self.logger.error(f"无法创建默认导出路径: {e}")
            self.config['export_path'] = os.path.expanduser("~/Documents")

    def save_config(self):
        """保存配置"""
        try:
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(self.config, f, ensure_ascii=False, indent=2)
        except Exception as e:
            self.logger.error(f"保存配置失败: {e}")

    def load_cumulative_data(self):
        """加载累计数据"""
        try:
            # 尝试脚本目录
            cumulative_file = os.path.join(os.path.dirname(__file__), 'cumulative_data.pkl')
            if not os.path.exists(cumulative_file):
                # 尝试用户目录
                cumulative_file = os.path.join(os.path.expanduser("~"), 'margin_cumulative_data.pkl')

            if os.path.exists(cumulative_file):
                with open(cumulative_file, 'rb') as f:
                    data = pickle.load(f)
                    self.monthly_data = data.get('monthly', {})
                    self.yearly_data = data.get('yearly', {})
                    self.logger.info(f"加载累计数据: {len(self.monthly_data)}个月度, {len(self.yearly_data)}个年度")
        except Exception as e:
            self.logger.error(f"加载累计数据失败: {e}")
            self.monthly_data = {}
            self.yearly_data = {}

    def save_cumulative_data(self):
        """保存累计数据"""
        try:
            data = {
                'monthly': self.monthly_data,
                'yearly': self.yearly_data
            }
            try:
                cumulative_file = os.path.join(os.path.dirname(__file__), 'cumulative_data.pkl')
            except:
                cumulative_file = os.path.join(os.path.expanduser("~"), 'margin_cumulative_data.pkl')

            with open(cumulative_file, 'wb') as f:
                pickle.dump(data, f)
            self.logger.info("累计数据已保存")
        except Exception as e:
            self.logger.error(f"保存累计数据失败: {e}")

    def load_history_from_disk(self):
        """从磁盘加载历史采购数据"""
        if os.path.exists(self.history_data_file) and self.config.get('remember_history', True):
            try:
                with open(self.history_data_file, 'rb') as f:
                    self.history_purchase_data = pickle.load(f)
                self.history_loaded = True
                if self.history_purchase_data is not None:
                    record_count = len(self.history_purchase_data)
                    unique_products = self.history_purchase_data['商品编码'].nunique()
                    history_info = f"已加载历史数据: {record_count}条记录, {unique_products}个商品"
                    self.history_file_path.set(history_info)
                    self.logger.info(f"加载历史采购数据: {record_count}条记录")
                else:
                    self.history_file_path.set("历史采购数据已加载（空数据）")
            except Exception as e:
                self.logger.error(f"加载历史数据失败: {e}")
                self.history_purchase_data = None
                self.history_loaded = False
                self.history_file_path.set("历史采购数据未加载")
        else:
            self.history_purchase_data = None
            self.history_loaded = False
            self.history_file_path.set("历史采购数据未加载")

    def save_history_to_disk(self):
        """保存历史采购数据到磁盘"""
        if self.history_purchase_data is not None and self.config.get('remember_history', True):
            try:
                with open(self.history_data_file, 'wb') as f:
                    pickle.dump(self.history_purchase_data, f)
                self.logger.info("历史采购数据已保存到磁盘")
            except Exception as e:
                self.logger.error(f"保存历史数据失败: {e}")

    def clear_history_data(self):
        """清空历史采购数据"""
        self.history_purchase_data = None
        self.history_loaded = False
        self.history_file_path.set("历史采购数据未加载")
        if os.path.exists(self.history_data_file):
            try:
                os.remove(self.history_data_file)
                self.logger.info("历史数据文件已删除")
            except Exception as e:
                self.logger.error(f"删除历史数据文件失败: {e}")

    def setup_menu(self):
        """设置菜单栏"""
        menubar = tk.Menu(self.root)
        self.root.config(menu=menubar)

        # 在 macOS 上，应用菜单会自动出现在顶部，这里定义的菜单会作为子菜单
        file_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="文件", menu=file_menu)
        file_menu.add_command(label="设置", command=self.open_settings)
        file_menu.add_separator()
        file_menu.add_command(label="导出历史数据", command=self.export_history_data)
        file_menu.add_command(label="清除历史数据", command=self.clear_history_data)
        file_menu.add_separator()
        file_menu.add_command(label="退出", command=self.root.quit)

        tools_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="工具", menu=tools_menu)
        tools_menu.add_command(label="生成模板文件", command=self.generate_templates)
        tools_menu.add_command(label="检查数据格式", command=self.check_data_format)
        tools_menu.add_command(label="手动加载历史数据", command=self.load_history_data)
        tools_menu.add_command(label="查看累计数据", command=self.view_cumulative_data)

        help_menu = tk.Menu(menubar, tearoff=0)
        menubar.add_cascade(label="帮助", menu=help_menu)
        help_menu.add_command(label="使用指南", command=self.show_guide)
        help_menu.add_command(label="版本信息", command=self.show_version)

    def setup_ui(self):
        """设置现代化用户界面"""
        main_container = tk.Frame(self.root, bg=self.colors['lighter'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=15, pady=15)

        # 标题栏
        self.setup_title_bar(main_container)

        self.paned = tk.PanedWindow(main_container, orient=tk.HORIZONTAL, sashrelief='raised', bg=self.colors['border'])
        self.paned.pack(fill=tk.BOTH, expand=True, pady=(15, 0))

        left_panel = tk.Frame(self.paned, bg=self.colors['white'], relief='flat', bd=1)
        self.setup_left_panel(left_panel)
        self.paned.add(left_panel, minsize=400)

        right_panel = tk.Frame(self.paned, bg=self.colors['white'], relief='flat', bd=1)
        self.setup_right_panel(right_panel)
        self.paned.add(right_panel, minsize=900)

        self.setup_status_bar(main_container)

    def setup_title_bar(self, parent):
        """设置标题栏"""
        title_frame = tk.Frame(parent, bg=self.colors['primary'], height=80)
        title_frame.pack(fill=tk.X)
        title_frame.pack_propagate(False)

        tk.Label(title_frame,
                 text="销售毛利智能分析系统",
                 font=(self.default_font, 20, 'bold'),
                 fg=self.colors['white'],
                 bg=self.colors['primary']).pack(side=tk.LEFT, padx=25, pady=10)

        info_frame = tk.Frame(title_frame, bg=self.colors['primary'])
        info_frame.pack(side=tk.RIGHT, padx=25, pady=10)

        self.date_label = tk.Label(info_frame,
                                   text=f"系统日期: {datetime.now().strftime('%Y年%m月%d日')}",
                                   font=(self.default_font, 10),
                                   fg=self.colors['white'],
                                   bg=self.colors['primary'])
        self.date_label.pack(side=tk.RIGHT, padx=(20, 0))

        tk.Label(info_frame,
                 text="专业版 v4.4 (Mac)",
                 font=(self.default_font, 9),
                 fg='#cccccc',
                 bg=self.colors['primary']).pack(side=tk.RIGHT)

    def setup_left_panel(self, parent):
        """设置左侧控制面板"""
        notebook = ttk.Notebook(parent)
        notebook.pack(fill=tk.BOTH, expand=True, padx=1, pady=1)

        data_tab = tk.Frame(notebook, bg=self.colors['white'])
        notebook.add(data_tab, text="📁 数据导入")
        self.setup_data_tab(data_tab)

        settings_tab = tk.Frame(notebook, bg=self.colors['white'])
        notebook.add(settings_tab, text="⚙️ 分析设置")
        self.setup_settings_tab(settings_tab)

    def setup_data_tab(self, parent):
        """设置数据导入标签页"""
        content = tk.Frame(parent, bg=self.colors['white'], padx=20, pady=20)
        content.pack(fill=tk.BOTH, expand=True)

        tk.Label(content,
                 text="历史采购数据 (自动更新)",
                 font=(self.default_font, 12, 'bold'),
                 fg=self.colors['dark'],
                 bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 5))

        history_frame = tk.Frame(content, bg=self.colors['white'])
        history_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Label(history_frame,
                 textvariable=self.history_file_path,
                 font=(self.default_font, 10),
                 fg=self.colors['dark'],
                 bg=self.colors['white'],
                 wraplength=320).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        history_btn_frame = tk.Frame(history_frame, bg=self.colors['white'])
        history_btn_frame.pack(side=tk.RIGHT)

        tk.Button(history_btn_frame,
                  text="手动加载",
                  command=self.load_history_data,
                  font=(self.default_font, 9),
                  bg=self.colors['accent'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=12,
                  pady=4).pack(side=tk.LEFT, padx=(0, 5))

        tk.Button(history_btn_frame,
                  text="清空",
                  command=self.clear_history_data,
                  font=(self.default_font, 9),
                  bg=self.colors['warning'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=12,
                  pady=4).pack(side=tk.LEFT)

        tk.Label(content,
                 text="销售数据表 (每日)",
                 font=(self.default_font, 12, 'bold'),
                 fg=self.colors['dark'],
                 bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 5))

        sales_frame = tk.Frame(content, bg=self.colors['white'])
        sales_frame.pack(fill=tk.X, pady=(0, 20))

        tk.Entry(sales_frame,
                 textvariable=self.sales_file_path,
                 font=(self.default_font, 10),
                 width=40,
                 relief='solid',
                 bd=1).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        tk.Button(sales_frame,
                  text="选择文件",
                  command=self.select_sales_file,
                  font=(self.default_font, 9),
                  bg=self.colors['primary'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=15,
                  pady=6).pack(side=tk.RIGHT)

        tk.Label(content,
                 text="最新采购价格表 (每日)",
                 font=(self.default_font, 12, 'bold'),
                 fg=self.colors['dark'],
                 bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 5))

        latest_frame = tk.Frame(content, bg=self.colors['white'])
        latest_frame.pack(fill=tk.X, pady=(0, 30))

        tk.Entry(latest_frame,
                 textvariable=self.latest_purchase_file_path,
                 font=(self.default_font, 10),
                 width=40,
                 relief='solid',
                 bd=1).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        tk.Button(latest_frame,
                  text="选择文件",
                  command=self.select_latest_purchase_file,
                  font=(self.default_font, 9),
                  bg=self.colors['primary'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=15,
                  pady=6).pack(side=tk.RIGHT)

        tk.Button(content,
                  text="📊 预览数据格式",
                  command=self.preview_data,
                  font=(self.default_font, 10),
                  bg=self.colors['accent'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=20,
                  pady=8).pack(fill=tk.X, pady=(10, 0))

        tk.Button(content,
                  text="🚀 开始分析计算",
                  command=self.calculate_margin,
                  font=(self.default_font, 13, 'bold'),
                  bg=self.colors['success'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=30,
                  pady=15,
                  cursor='hand2').pack(fill=tk.X, pady=(20, 0))

        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(content,
                                            variable=self.progress_var,
                                            maximum=100,
                                            mode='determinate',
                                            length=300)
        self.progress_bar.pack(fill=tk.X, pady=(20, 5))

        self.status_label = tk.Label(content,
                                     text="就绪",
                                     font=(self.default_font, 9),
                                     fg=self.colors['dark'],
                                     bg=self.colors['white'])
        self.status_label.pack()

    def setup_settings_tab(self, parent):
        """设置分析设置标签页"""
        content = tk.Frame(parent, bg=self.colors['white'], padx=20, pady=20)
        content.pack(fill=tk.BOTH, expand=True)

        tk.Label(content,
                 text="分析时间范围",
                 font=(self.default_font, 12, 'bold'),
                 fg=self.colors['dark'],
                 bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 10))

        time_frame = tk.Frame(content, bg=self.colors['white'])
        time_frame.pack(fill=tk.X, pady=(0, 20))

        self.analysis_type = tk.StringVar(value="daily")

        tk.Radiobutton(time_frame,
                       text="当日分析",
                       variable=self.analysis_type,
                       value="daily",
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W)

        tk.Radiobutton(time_frame,
                       text="月度累计",
                       variable=self.analysis_type,
                       value="monthly",
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=(5, 0))

        tk.Radiobutton(time_frame,
                       text="年度累计",
                       variable=self.analysis_type,
                       value="yearly",
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=(5, 0))

        tk.Label(content,
                 text="导出选项",
                 font=(self.default_font, 12, 'bold'),
                 fg=self.colors['dark'],
                 bg=self.colors['white']).pack(anchor=tk.W, pady=(10, 10))

        self.auto_open_var = tk.BooleanVar(value=self.config.get('auto_open', True))
        tk.Checkbutton(content,
                       text="分析完成后自动打开Excel文件",
                       variable=self.auto_open_var,
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 8))

        self.create_charts_var = tk.BooleanVar(value=True)
        tk.Checkbutton(content,
                       text="生成分析图表",
                       variable=self.create_charts_var,
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 8))

        self.remember_history_var = tk.BooleanVar(value=self.config.get('remember_history', True))
        tk.Checkbutton(content,
                       text="记住历史采购数据",
                       variable=self.remember_history_var,
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 8))

        tk.Label(content,
                 text="导出路径",
                 font=(self.default_font, 10, 'bold'),
                 fg=self.colors['dark'],
                 bg=self.colors['white']).pack(anchor=tk.W, pady=(10, 5))

        path_frame = tk.Frame(content, bg=self.colors['white'])
        path_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Button(path_frame,
                  text="修改导出路径",
                  command=self.change_export_path,
                  font=(self.default_font, 9),
                  bg=self.colors['warning'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=12,
                  pady=4).pack(side=tk.LEFT)

        tk.Label(path_frame,
                 text=self.config['export_path'],
                 font=(self.default_font, 9),
                 fg=self.colors['dark'],
                 bg=self.colors['white'],
                 wraplength=300,
                 justify=tk.LEFT).pack(side=tk.LEFT, padx=(10, 0), fill=tk.X, expand=True)

        tk.Button(content,
                  text="🔄 清空所有数据",
                  command=self.reset_all,
                  font=(self.default_font, 11),
                  bg=self.colors['light'],
                  fg=self.colors['dark'],
                  relief='flat',
                  padx=20,
                  pady=10).pack(fill=tk.X, pady=(20, 5))

        tk.Button(content,
                  text="📂 打开导出文件夹",
                  command=self.open_export_folder,
                  font=(self.default_font, 11),
                  bg=self.colors['light'],
                  fg=self.colors['dark'],
                  relief='flat',
                  padx=20,
                  pady=10).pack(fill=tk.X, pady=(5, 5))

        tk.Button(content,
                  text="📊 数据合并与更新",
                  command=self.merge_and_update_data,
                  font=(self.default_font, 11),
                  bg=self.colors['accent'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=20,
                  pady=10).pack(fill=tk.X, pady=(5, 0))

    def setup_right_panel(self, parent):
        """设置右侧显示面板"""
        style = ttk.Style()
        style.theme_use('clam')  # 使用 'clam' 主题以便跨平台自定义颜色

        self.notebook = ttk.Notebook(parent)
        self.notebook.pack(fill=tk.BOTH, expand=True, padx=2, pady=2)

        self.summary_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(self.summary_frame, text="📊 汇总分析")
        self.setup_summary_tab()

        self.chart_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(self.chart_frame, text="📈 图表展示")
        self.setup_chart_tab()

        self.detail_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(self.detail_frame, text="📋 详细数据")
        self.setup_detail_tab()

        self.history_data_frame = tk.Frame(self.notebook, bg=self.colors['white'])
        self.notebook.add(self.history_data_frame, text="📚 历史数据")
        self.setup_history_data_tab()

    def setup_detail_tab(self):
        """设置详细数据标签页"""
        columns = ('序号', '商品编码', '商品名称', '门店名称', '一级分类', '二级分类',
                   '订货数量', '商品单价（元）', '销售金额（元）', '采购单价（元）', '采购成本（元）',
                   '销售毛利（元）', '毛利率')

        tree_container = tk.Frame(self.detail_frame, bg=self.colors['white'])
        tree_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        search_frame = tk.Frame(tree_container, bg=self.colors['white'])
        search_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(search_frame,
                 text="搜索:",
                 font=(self.default_font, 10),
                 bg=self.colors['white']).pack(side=tk.LEFT, padx=(0, 10))

        self.search_var = tk.StringVar()
        self.search_entry = tk.Entry(search_frame,
                                     textvariable=self.search_var,
                                     font=(self.default_font, 10),
                                     width=40)
        self.search_entry.pack(side=tk.LEFT)
        self.search_entry.bind('<Return>', lambda event: self.search_data())

        tk.Button(search_frame,
                  text="搜索",
                  command=self.search_data,
                  font=(self.default_font, 9),
                  bg=self.colors['primary'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=15,
                  pady=4).pack(side=tk.LEFT, padx=(10, 0))

        tk.Button(search_frame,
                  text="清空",
                  command=self.clear_search,
                  font=(self.default_font, 9),
                  bg=self.colors['warning'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=15,
                  pady=4).pack(side=tk.LEFT, padx=(5, 0))

        self.tree = ttk.Treeview(tree_container, columns=columns, show='headings', height=25)

        col_widths = [60, 100, 150, 100, 90, 90, 90, 90, 100, 90, 100, 100, 80]
        for col, width in zip(columns, col_widths):
            self.tree.heading(col, text=col)
            self.tree.column(col, width=width, anchor='center')

        scrollbar = ttk.Scrollbar(tree_container, orient="vertical", command=self.tree.yview)
        self.tree.configure(yscrollcommand=scrollbar.set)

        self.tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

    def setup_summary_tab(self):
        """设置汇总分析标签页"""
        self.summary_text = tk.Text(self.summary_frame,
                                    font=(self.default_font, 10),
                                    bg=self.colors['white'],
                                    fg=self.colors['dark'],
                                    wrap='word',
                                    padx=20,
                                    pady=20)

        scrollbar = ttk.Scrollbar(self.summary_frame)
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        self.summary_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        self.summary_text.config(yscrollcommand=scrollbar.set)
        scrollbar.config(command=self.summary_text.yview)

        self.summary_text.tag_configure('title', font=(self.default_font, 16, 'bold'),
                                        foreground=self.colors['primary'], spacing3=10)
        self.summary_text.tag_configure('section', font=(self.default_font, 13, 'bold'),
                                        foreground=self.colors['secondary'], spacing3=8)
        self.summary_text.tag_configure('subtitle', font=(self.default_font, 11, 'bold'),
                                        foreground=self.colors['dark'])
        self.summary_text.tag_configure('highlight', font=(self.default_font, 11, 'bold'),
                                        foreground=self.colors['success'])
        self.summary_text.tag_configure('warning', font=(self.default_font, 11, 'bold'),
                                        foreground=self.colors['danger'])
        self.summary_text.tag_configure('data', font=(self.default_font, 10))

    def setup_chart_tab(self):
        """设置图表展示标签页 - 四宫格布局"""
        self.chart_container = tk.Frame(self.chart_frame, bg=self.colors['white'])
        self.chart_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        self.chart_frame1 = tk.Frame(self.chart_container, bg=self.colors['white'],
                                     highlightbackground=self.colors['border'], highlightthickness=1)
        self.chart_frame2 = tk.Frame(self.chart_container, bg=self.colors['white'],
                                     highlightbackground=self.colors['border'], highlightthickness=1)
        self.chart_frame3 = tk.Frame(self.chart_container, bg=self.colors['white'],
                                     highlightbackground=self.colors['border'], highlightthickness=1)
        self.chart_frame4 = tk.Frame(self.chart_container, bg=self.colors['white'],
                                     highlightbackground=self.colors['border'], highlightthickness=1)

        self.chart_frame1.grid(row=0, column=0, sticky='nsew', padx=5, pady=5)
        self.chart_frame2.grid(row=0, column=1, sticky='nsew', padx=5, pady=5)
        self.chart_frame3.grid(row=1, column=0, sticky='nsew', padx=5, pady=5)
        self.chart_frame4.grid(row=1, column=1, sticky='nsew', padx=5, pady=5)

        self.chart_container.grid_rowconfigure(0, weight=1)
        self.chart_container.grid_rowconfigure(1, weight=1)
        self.chart_container.grid_columnconfigure(0, weight=1)
        self.chart_container.grid_columnconfigure(1, weight=1)

        self.chart_label1 = tk.Label(self.chart_frame1, text="销售金额占比", font=(self.default_font, 11, 'bold'),
                                     bg=self.colors['white'], fg=self.colors['dark'])
        self.chart_label1.pack(pady=5)

        self.chart_label2 = tk.Label(self.chart_frame2, text="销售毛利占比", font=(self.default_font, 11, 'bold'),
                                     bg=self.colors['white'], fg=self.colors['dark'])
        self.chart_label2.pack(pady=5)

        self.chart_label3 = tk.Label(self.chart_frame3, text="门店销售排行", font=(self.default_font, 11, 'bold'),
                                     bg=self.colors['white'], fg=self.colors['dark'])
        self.chart_label3.pack(pady=5)

        self.chart_label4 = tk.Label(self.chart_frame4, text="品类毛利排行", font=(self.default_font, 11, 'bold'),
                                     bg=self.colors['white'], fg=self.colors['dark'])
        self.chart_label4.pack(pady=5)

    def setup_history_data_tab(self):
        """设置历史数据标签页"""
        main_container = tk.Frame(self.history_data_frame, bg=self.colors['white'])
        main_container.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        search_frame = tk.Frame(main_container, bg=self.colors['white'])
        search_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Label(search_frame,
                 text="搜索商品:",
                 font=(self.default_font, 10),
                 bg=self.colors['white']).pack(side=tk.LEFT, padx=(0, 10))

        self.history_search_var = tk.StringVar()
        self.history_search_entry = tk.Entry(search_frame,
                                             textvariable=self.history_search_var,
                                             font=(self.default_font, 10),
                                             width=30)
        self.history_search_entry.pack(side=tk.LEFT)
        self.history_search_entry.bind('<Return>', lambda event: self.search_history_data())

        tk.Button(search_frame,
                  text="搜索",
                  command=self.search_history_data,
                  font=(self.default_font, 9),
                  bg=self.colors['primary'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=15,
                  pady=4).pack(side=tk.LEFT, padx=(10, 0))

        tk.Button(search_frame,
                  text="生成价格趋势图",
                  command=self.generate_price_trend,
                  font=(self.default_font, 9),
                  bg=self.colors['success'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=15,
                  pady=4).pack(side=tk.LEFT, padx=(5, 0))

        tk.Button(search_frame,
                  text="清空搜索",
                  command=self.clear_history_search,
                  font=(self.default_font, 9),
                  bg=self.colors['warning'],
                  fg=self.colors['white'],
                  relief='flat',
                  padx=15,
                  pady=4).pack(side=tk.LEFT, padx=(5, 0))

        paned = tk.PanedWindow(main_container, orient=tk.HORIZONTAL, sashrelief='raised', bg=self.colors['border'])
        paned.pack(fill=tk.BOTH, expand=True)

        left_frame = tk.Frame(paned, bg=self.colors['white'])
        paned.add(left_frame, minsize=500)

        self.history_info_label = tk.Label(left_frame,
                                           text="历史采购数据未加载",
                                           font=(self.default_font, 11),
                                           fg=self.colors['dark'],
                                           bg=self.colors['white'])
        self.history_info_label.pack(anchor=tk.W, pady=(0, 10), padx=10)

        columns = ('商品编码', '商品名称', '采购单价（元）', '建单时间')
        self.history_tree = ttk.Treeview(left_frame, columns=columns, show='headings', height=20)

        col_widths = [100, 150, 100, 120]
        for col, width in zip(columns, col_widths):
            self.history_tree.heading(col, text=col)
            self.history_tree.column(col, width=width, anchor='center')

        scrollbar = ttk.Scrollbar(left_frame, orient="vertical", command=self.history_tree.yview)
        self.history_tree.configure(yscrollcommand=scrollbar.set)

        self.history_tree.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(10, 0))
        scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

        right_frame = tk.Frame(paned, bg=self.colors['white'])
        paned.add(right_frame, minsize=500)

        self.trend_chart_label = tk.Label(right_frame,
                                          text="价格趋势图",
                                          font=(self.default_font, 12, 'bold'),
                                          fg=self.colors['dark'],
                                          bg=self.colors['white'])
        self.trend_chart_label.pack(pady=(10, 5))

        self.trend_info_label = tk.Label(right_frame,
                                         text="请先搜索商品查看价格趋势",
                                         font=(self.default_font, 10),
                                         fg=self.colors['dark'],
                                         bg=self.colors['white'])
        self.trend_info_label.pack(pady=(0, 10))

        self.trend_canvas_frame = tk.Frame(right_frame, bg=self.colors['white'])
        self.trend_canvas_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

    def setup_status_bar(self, parent):
        """设置状态栏"""
        status_frame = tk.Frame(parent, bg=self.colors['primary'], height=30)
        status_frame.pack(fill=tk.X, pady=(15, 0))
        status_frame.pack_propagate(False)

        self.status_var = tk.StringVar(value="就绪")
        status_label = tk.Label(status_frame,
                                textvariable=self.status_var,
                                font=(self.default_font, 9),
                                fg=self.colors['white'],
                                bg=self.colors['primary'])
        status_label.pack(side=tk.LEFT, padx=20)

        self.history_status_var = tk.StringVar(value="历史数据: 未加载")
        history_label = tk.Label(status_frame,
                                 textvariable=self.history_status_var,
                                 font=(self.default_font, 9),
                                 fg=self.colors['white'],
                                 bg=self.colors['primary'])
        history_label.pack(side=tk.LEFT, padx=20)

    # ================= 文件选择方法 =================
    def select_sales_file(self):
        filename = filedialog.askopenfilename(
            title="选择销售数据表",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if filename:
            self.sales_file_path.set(filename)
            self.logger.info(f"已选择销售文件: {os.path.basename(filename)}")

    def select_latest_purchase_file(self):
        filename = filedialog.askopenfilename(
            title="选择最新采购价格表",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if filename:
            self.latest_purchase_file_path.set(filename)
            self.logger.info(f"已选择最新采购文件: {os.path.basename(filename)}")

    def load_history_data(self):
        """加载历史采购数据"""
        filename = filedialog.askopenfilename(
            title="选择历史采购价格表",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not filename:
            return

        try:
            self.update_progress(0, "正在加载历史采购数据...")
            history_data = pd.read_excel(filename)
            history_data = self.auto_rename_columns(history_data, 'purchase_history')

            required_cols = ['商品编码', '采购单价', '建单时间']
            missing_cols = [col for col in required_cols if col not in history_data.columns]
            if missing_cols:
                raise ValueError(f"历史采购数据缺少必要列: {', '.join(missing_cols)}")

            if not pd.api.types.is_datetime64_any_dtype(history_data['建单时间']):
                history_data['建单时间'] = pd.to_datetime(history_data['建单时间'], errors='coerce')

            self.history_purchase_data = history_data
            self.history_loaded = True

            record_count = len(history_data)
            unique_products = history_data['商品编码'].nunique()

            min_date = history_data['建单时间'].min()
            max_date = history_data['建单时间'].max()
            date_str = ""
            if pd.notna(min_date) and pd.notna(max_date):
                date_str = f", 时间范围: {min_date.strftime('%Y-%m-%d')} 至 {max_date.strftime('%Y-%m-%d')}"

            status_text = f"历史数据已加载: {record_count}条记录, {unique_products}个商品"
            self.history_file_path.set(status_text)
            self.history_status_var.set(f"历史数据: {record_count}条记录")

            self.save_history_to_disk()
            self.update_history_data_tab()

            self.update_progress(100, "历史采购数据加载完成！")
            self.logger.info(f"历史采购数据加载成功: {record_count}条记录")
            messagebox.showinfo("成功",
                                f"历史采购数据加载成功！\n\n记录数: {record_count}\n商品数: {unique_products}\n{date_str}")

        except Exception as e:
            self.logger.error(f"加载历史采购数据失败: {str(e)}")
            messagebox.showerror("错误", f"加载历史采购数据失败: {str(e)}")

    def export_history_data(self):
        """导出历史采购数据"""
        if self.history_purchase_data is None:
            messagebox.showwarning("警告", "没有可导出的历史数据")
            return

        filename = filedialog.asksaveasfilename(
            title="保存历史采购数据",
            defaultextension=".xlsx",
            filetypes=[("Excel文件", "*.xlsx"), ("CSV文件", "*.csv")]
        )
        if filename:
            try:
                if filename.endswith('.csv'):
                    self.history_purchase_data.to_csv(filename, index=False, encoding='utf-8-sig')
                else:
                    self.history_purchase_data.to_excel(filename, index=False)

                self.logger.info(f"历史数据已导出到: {filename}")
                messagebox.showinfo("成功", f"历史数据已成功导出到:\n{filename}")

            except Exception as e:
                self.logger.error(f"导出历史数据失败: {str(e)}")
                messagebox.showerror("错误", f"导出历史数据失败: {str(e)}")

    def update_history_data_tab(self):
        """更新历史数据标签页"""
        for item in self.history_tree.get_children():
            self.history_tree.delete(item)

        if self.history_purchase_data is None:
            self.history_info_label.config(text="历史采购数据未加载")
            return

        record_count = len(self.history_purchase_data)
        unique_products = self.history_purchase_data['商品编码'].nunique()
        min_date = self.history_purchase_data['建单时间'].min()
        max_date = self.history_purchase_data['建单时间'].max()

        info_text = f"历史采购数据: {record_count}条记录, {unique_products}个商品"
        if pd.notna(min_date) and pd.notna(max_date):
            info_text += f", 时间范围: {min_date.strftime('%Y-%m-%d')} 至 {max_date.strftime('%Y-%m-%d')}"

        self.history_info_label.config(text=info_text)

        display_data = self.history_purchase_data.head(200)

        for _, row in display_data.iterrows():
            values = (
                row.get('商品编码', ''),
                row.get('商品名称', ''),
                f"{row.get('采购单价', 0):,.2f}",
                row.get('建单时间', '')
            )
            self.history_tree.insert('', tk.END, values=values)

    def change_export_path(self):
        """修改导出路径"""
        path = filedialog.askdirectory(title="选择导出文件夹")
        if path:
            self.config['export_path'] = path
            self.save_config()
            self.logger.info(f"导出路径已更新为: {path}")
            messagebox.showinfo("成功", f"导出路径已更新为:\n{path}")

    def open_settings(self):
        """打开设置窗口"""
        settings_window = tk.Toplevel(self.root)
        settings_window.title("系统设置")
        settings_window.geometry("500x400")
        settings_window.configure(bg=self.colors['lighter'])
        settings_window.transient(self.root)
        settings_window.grab_set()

        settings_window.update_idletasks()
        width = settings_window.winfo_width()
        height = settings_window.winfo_height()
        x = (settings_window.winfo_screenwidth() // 2) - (width // 2)
        y = (settings_window.winfo_screenheight() // 2) - (height // 2)
        settings_window.geometry(f'{width}x{height}+{x}+{y}')

        tk.Label(settings_window,
                 text="⚙️ 系统设置",
                 font=(self.default_font, 16, 'bold'),
                 fg=self.colors['primary'],
                 bg=self.colors['lighter']).pack(pady=(20, 10))

        export_frame = tk.Frame(settings_window, bg=self.colors['white'], padx=20, pady=15)
        export_frame.pack(fill=tk.X, padx=20, pady=10)

        tk.Label(export_frame,
                 text="导出文件路径",
                 font=(self.default_font, 11, 'bold'),
                 fg=self.colors['dark'],
                 bg=self.colors['white']).pack(anchor=tk.W, pady=(0, 5))

        export_path_var = tk.StringVar(value=self.config['export_path'])
        entry_frame = tk.Frame(export_frame, bg=self.colors['white'])
        entry_frame.pack(fill=tk.X, pady=(0, 10))

        tk.Entry(entry_frame,
                 textvariable=export_path_var,
                 font=(self.default_font, 9),
                 width=50).pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(0, 10))

        def select_export_path():
            path = filedialog.askdirectory(title="选择导出文件夹")
            if path:
                export_path_var.set(path)

        tk.Button(entry_frame,
                  text="选择文件夹",
                  command=select_export_path,
                  font=(self.default_font, 9),
                  bg=self.colors['accent'],
                  fg=self.colors['white']).pack(side=tk.RIGHT)

        other_frame = tk.Frame(settings_window, bg=self.colors['white'], padx=20, pady=15)
        other_frame.pack(fill=tk.X, padx=20, pady=10)

        auto_open_var = tk.BooleanVar(value=self.config.get('auto_open', True))
        tk.Checkbutton(other_frame,
                       text="分析完成后自动打开Excel文件",
                       variable=auto_open_var,
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=5)

        remember_history_var = tk.BooleanVar(value=self.config.get('remember_history', True))
        tk.Checkbutton(other_frame,
                       text="记住历史采购数据",
                       variable=remember_history_var,
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=5)

        create_charts_var = tk.BooleanVar(value=self.config.get('create_charts', True))
        tk.Checkbutton(other_frame,
                       text="生成分析图表",
                       variable=create_charts_var,
                       font=(self.default_font, 10),
                       bg=self.colors['white']).pack(anchor=tk.W, pady=5)

        def save_settings():
            self.config['export_path'] = export_path_var.get()
            self.config['auto_open'] = auto_open_var.get()
            self.config['remember_history'] = remember_history_var.get()
            self.config['create_charts'] = create_charts_var.get()
            self.save_config()

            self.create_charts_var.set(create_charts_var.get())
            self.remember_history_var.set(remember_history_var.get())

            if not remember_history_var.get() and os.path.exists(self.history_data_file):
                try:
                    os.remove(self.history_data_file)
                except:
                    pass

            self.logger.info("系统设置已保存")
            messagebox.showinfo("成功", "设置已保存")
            settings_window.destroy()

        button_frame = tk.Frame(settings_window, bg=self.colors['lighter'])
        button_frame.pack(fill=tk.X, padx=20, pady=20)

        tk.Button(button_frame,
                  text="💾 保存设置",
                  command=save_settings,
                  font=(self.default_font, 11, 'bold'),
                  bg=self.colors['success'],
                  fg=self.colors['white'],
                  padx=30,
                  pady=10).pack()

    def generate_templates(self):
        """生成模板文件"""
        template_dir = filedialog.askdirectory(title="选择模板保存位置")
        if template_dir:
            sales_template = pd.DataFrame({
                '商品编码': ['SP001', 'SP002', 'SP003'],
                '商品名称': ['商品A', '商品B', '商品C'],
                '门店名称': ['门店1', '门店1', '门店2'],
                '一级分类': ['分类1', '分类1', '分类2'],
                '二级分类': ['子类1', '子类2', '子类1'],
                '订货数量': [10, 20, 15],
                '商品单价': [100.0, 50.0, 80.0],
                '销售日期': [datetime.now()] * 3
            })

            purchase_template = pd.DataFrame({
                '商品编码': ['SP001', 'SP002', 'SP003'],
                '商品名称': ['商品A', '商品B', '商品C'],
                '采购单价': [60.0, 30.0, 50.0],
                '建单时间': [datetime.now()] * 3
            })

            try:
                sales_template.to_excel(os.path.join(template_dir, "销售数据模板.xlsx"), index=False)
                purchase_template.to_excel(os.path.join(template_dir, "采购数据模板.xlsx"), index=False)

                self.logger.info(f"模板文件已生成到: {template_dir}")
                messagebox.showinfo("成功", f"模板文件已生成到:\n{template_dir}")
            except Exception as e:
                messagebox.showerror("错误", f"生成模板失败: {str(e)}")

    def check_data_format(self):
        """检查数据格式"""
        if not self.sales_file_path.get():
            messagebox.showwarning("警告", "请先选择销售数据文件")
            return

        try:
            df = pd.read_excel(self.sales_file_path.get(), nrows=5)
            required_cols = ['商品编码', '商品名称', '订货数量', '商品单价']
            missing = [col for col in required_cols if col not in df.columns]

            if missing:
                messagebox.showwarning("格式检查", f"缺少必要列: {', '.join(missing)}")
            else:
                messagebox.showinfo("格式检查", "数据格式正确！")

        except Exception as e:
            self.logger.error(f"检查数据格式失败: {str(e)}")
            messagebox.showerror("错误", f"读取文件失败: {str(e)}")

    def refresh_ui(self):
        """刷新界面"""
        self.date_label.config(text=f"系统日期: {datetime.now().strftime('%Y年%m月%d日')}")

    def show_guide(self):
        """显示使用指南"""
        guide = """使用指南：

1. 初始化设置：
   - 第一次使用时，点击"工具->手动加载历史数据"加载历史采购数据
   - 或者直接开始日常分析，系统会自动累计采购数据

2. 日常使用：
   - 每日选择销售数据表
   - 每日选择最新采购价格表
   - 点击"开始分析计算"
   - 系统会自动将最新采购数据追加到历史库中

3. 数据更新：
   - 当有新商品或价格变动时，通过最新采购价格表更新
   - 如需完全更新历史数据，可重新手动加载历史采购数据

4. 分析设置：
   - 可选择当日、月度、年度分析
   - 月度/年度分析会自动累加历史累计数据
   - 可设置是否自动打开结果文件
   - 可自定义导出路径

5. 注意事项：
   - 历史采购数据保存在程序内部，重启后仍可用
   - 最新采购数据优先于历史数据
   - 缺失的采购价格会自动从历史数据获取
   - 文件夹按月自动生成"""

        messagebox.showinfo("使用指南", guide)

    def show_version(self):
        """显示版本信息"""
        version = """销售毛利分析系统 v4.4 (Mac适配版)

新功能特性：
✓ 历史采购数据自动追加（无需手动合并）
✓ 月度/年度累计数据自动累加
✓ 导出报表按销售最新日期命名
✓ 文件夹按销售月份自动归档
✓ 新增每日趋势分析工作表
✓ Excel毛利率格式修正（26.33%）
✓ 智能价格合并机制
✓ 四宫格图表展示
✓ 历史数据价格趋势分析

Mac优化：
✓ 自动适配 PingFang SC 中文字体
✓ 优化文件打开方式
✓ 兼容 Retina 显示屏渲染

历史数据管理：
- 通过"工具->手动加载历史数据"加载
- 通过"文件->清除历史数据"删除
- 历史数据重启后仍可用

© 2025 数据智能分析平台
作者：NASHH"""

        messagebox.showinfo("版本信息", version)

    def preview_data(self):
        """预览数据格式"""
        files = [
            ("销售数据", self.sales_file_path.get()),
            ("最新采购", self.latest_purchase_file_path.get())
        ]

        preview_text = "数据格式预览:\n\n"

        for name, path in files:
            if path:
                try:
                    df = pd.read_excel(path, nrows=3)
                    preview_text += f"{name}表:\n"
                    preview_text += f"  行数: {len(df)}, 列数: {len(df.columns)}\n"
                    preview_text += f"  列名: {', '.join(df.columns)}\n\n"
                except Exception as e:
                    preview_text += f"{name}表: 读取失败 ({str(e)})\n\n"
            else:
                preview_text += f"{name}表: 未选择文件\n\n"

        if self.history_purchase_data is not None:
            preview_text += "历史采购数据:\n"
            preview_text += f"  记录数: {len(self.history_purchase_data)}\n"
            preview_text += f"  商品数: {self.history_purchase_data['商品编码'].nunique()}\n"
            if '建单时间' in self.history_purchase_data.columns:
                min_date = self.history_purchase_data['建单时间'].min()
                max_date = self.history_purchase_data['建单时间'].max()
                preview_text += f"  时间范围: {min_date.strftime('%Y-%m-%d')} 至 {max_date.strftime('%Y-%m-%d')}\n\n"
        else:
            preview_text += "历史采购数据: 未加载\n\n"

        self.logger.info("数据格式预览完成")
        messagebox.showinfo("数据预览", preview_text)

    def reset_all(self):
        """清空所有数据"""
        self.sales_file_path.set("")
        self.latest_purchase_file_path.set("")
        self.clear_treeview()
        self.summary_text.delete(1.0, tk.END)
        self.progress_var.set(0)
        self.status_label.config(text="已重置")
        self.status_var.set("已重置")
        self.logger.info("所有数据已清空")

        for widget in self.chart_container.winfo_children():
            if isinstance(widget, tk.Frame) and widget not in [self.chart_frame1, self.chart_frame2, self.chart_frame3,
                                                               self.chart_frame4]:
                for sub_widget in widget.winfo_children():
                    sub_widget.destroy()
            elif not isinstance(widget, tk.Label):
                widget.destroy()

        self.clear_history_search()

    def open_export_folder(self):
        """打开导出文件夹"""
        path = self.config['export_path']
        if os.path.exists(path):
            if sys.platform == 'darwin':  # macOS
                import subprocess
                subprocess.run(['open', path])
            elif os.name == 'nt':  # Windows
                os.startfile(path)
            else:  # Linux
                import subprocess
                subprocess.run(['xdg-open', path])
        else:
            self.logger.warning("导出文件夹不存在")
            messagebox.showwarning("警告", "导出文件夹不存在")

    def view_cumulative_data(self):
        """查看累计数据"""
        cumulative_window = tk.Toplevel(self.root)
        cumulative_window.title("累计数据查看")
        cumulative_window.geometry("800x600")
        cumulative_window.configure(bg=self.colors['lighter'])
        cumulative_window.transient(self.root)
        cumulative_window.grab_set()

        cumulative_window.update_idletasks()
        width = cumulative_window.winfo_width()
        height = cumulative_window.winfo_height()
        x = (cumulative_window.winfo_screenwidth() // 2) - (width // 2)
        y = (cumulative_window.winfo_screenheight() // 2) - (height // 2)
        cumulative_window.geometry(f'{width}x{height}+{x}+{y}')

        notebook = ttk.Notebook(cumulative_window)
        notebook.pack(fill=tk.BOTH, expand=True, padx=10, pady=10)

        monthly_frame = tk.Frame(notebook, bg=self.colors['white'])
        notebook.add(monthly_frame, text="月度累计")

        if self.monthly_data:
            monthly_text = tk.Text(monthly_frame, font=(self.default_font, 10), wrap='word')
            scrollbar = ttk.Scrollbar(monthly_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            monthly_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            monthly_text.config(yscrollcommand=scrollbar.set)
            scrollbar.config(command=monthly_text.yview)

            monthly_text.insert(tk.END, "月度累计数据汇总\n\n", 'title')
            for month, data in self.monthly_data.items():
                monthly_text.insert(tk.END, f"{month}:\n", 'section')
                monthly_text.insert(tk.END, f"  总销售额: ¥{data.get('total_sales', 0):,.2f}\n")
                monthly_text.insert(tk.END, f"  总毛利: ¥{data.get('total_margin', 0):,.2f}\n")
                monthly_text.insert(tk.END, f"  毛利率: {data.get('margin_rate', 0):.2f}%\n")
                monthly_text.insert(tk.END, f"  商品种类: {data.get('product_count', 0)}\n\n")
        else:
            tk.Label(monthly_frame, text="暂无月度累计数据", font=(self.default_font, 12),
                     bg=self.colors['white']).pack(pady=50)

        yearly_frame = tk.Frame(notebook, bg=self.colors['white'])
        notebook.add(yearly_frame, text="年度累计")

        if self.yearly_data:
            yearly_text = tk.Text(yearly_frame, font=(self.default_font, 10), wrap='word')
            scrollbar = ttk.Scrollbar(yearly_frame)
            scrollbar.pack(side=tk.RIGHT, fill=tk.Y)
            yearly_text.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
            yearly_text.config(yscrollcommand=scrollbar.set)
            scrollbar.config(command=yearly_text.yview)

            yearly_text.insert(tk.END, "年度累计数据汇总\n\n", 'title')
            for year, data in self.yearly_data.items():
                yearly_text.insert(tk.END, f"{year}:\n", 'section')
                yearly_text.insert(tk.END, f"  总销售额: ¥{data.get('total_sales', 0):,.2f}\n")
                yearly_text.insert(tk.END, f"  总毛利: ¥{data.get('total_margin', 0):,.2f}\n")
                yearly_text.insert(tk.END, f"  毛利率: {data.get('margin_rate', 0):.2f}%\n")
                yearly_text.insert(tk.END, f"  商品种类: {data.get('product_count', 0)}\n\n")
        else:
            tk.Label(yearly_frame, text="暂无年度累计数据", font=(self.default_font, 12),
                     bg=self.colors['white']).pack(pady=50)

    def merge_and_update_data(self):
        """数据合并与更新"""
        if self.history_purchase_data is None:
            messagebox.showwarning("警告", "请先加载历史采购数据")
            return

        choice = messagebox.askyesno("数据合并",
                                     "是否将最新采购数据合并到历史数据中？\n\n"
                                     "这将更新历史采购数据，包括：\n"
                                     "1. 新增商品\n"
                                     "2. 更新价格\n"
                                     "3. 添加新的建单记录")

        if not choice:
            return

        filename = filedialog.askopenfilename(
            title="选择要合并的最新采购数据",
            filetypes=[("Excel文件", "*.xlsx *.xls"), ("CSV文件", "*.csv"), ("所有文件", "*.*")]
        )
        if not filename:
            return

        try:
            latest_data = pd.read_excel(filename)
            latest_data = self.auto_rename_columns(latest_data, 'purchase_latest')
            latest_data['建单时间'] = datetime.now()

            if '商品名称' not in latest_data.columns and '商品名称' in self.history_purchase_data.columns:
                name_mapping = self.history_purchase_data[['商品编码', '商品名称']].drop_duplicates('商品编码',
                                                                                                    keep='first')
                latest_data = pd.merge(latest_data, name_mapping, on='商品编码', how='left')

            combined_data = pd.concat([self.history_purchase_data, latest_data], ignore_index=True)
            combined_data = combined_data.sort_values('建单时间', ascending=False)
            combined_data = combined_data.drop_duplicates(['商品编码', '建单时间'], keep='first')

            self.history_purchase_data = combined_data
            self.save_history_to_disk()

            record_count = len(combined_data)
            self.history_file_path.set(f"历史数据已更新: {record_count}条记录")
            self.update_history_data_tab()

            self.logger.info(f"历史数据已更新，新增{len(latest_data)}条记录")
            messagebox.showinfo("成功",
                                f"历史数据更新成功！\n新增记录: {len(latest_data)}条\n总记录数: {record_count}条")

        except Exception as e:
            self.logger.error(f"合并数据失败: {str(e)}")
            messagebox.showerror("错误", f"合并数据失败: {str(e)}")

    def search_data(self):
        """搜索详细数据"""
        if self.current_detailed_data is None:
            messagebox.showwarning("警告", "请先运行分析计算")
            return

        keyword = self.search_var.get().strip()
        if not keyword:
            self.display_detailed_data(self.current_detailed_data)
            return

        try:
            data = self.current_detailed_data.copy()
            data['商品编码_str'] = data['商品编码'].astype(str)
            data['订货数量_str'] = data['订货数量'].astype(str)

            mask = (
                    data['商品编码_str'].str.contains(keyword, case=False, na=False) |
                    data['商品名称'].astype(str).str.contains(keyword, case=False, na=False) |
                    data['门店名称'].astype(str).str.contains(keyword, case=False, na=False) |
                    data['一级分类'].astype(str).str.contains(keyword, case=False, na=False) |
                    data['二级分类'].astype(str).str.contains(keyword, case=False, na=False) |
                    data['订货数量_str'].str.contains(keyword, case=False, na=False)
            )

            filtered_data = data[mask]

            if len(filtered_data) == 0:
                messagebox.showinfo("搜索结果", f"未找到包含 '{keyword}' 的记录")
                return

            self.display_detailed_data(filtered_data)
            self.status_var.set(f"找到 {len(filtered_data)} 条匹配记录")

        except Exception as e:
            self.logger.error(f"搜索数据失败: {str(e)}")
            messagebox.showerror("搜索错误", f"搜索失败: {str(e)}")

    def clear_search(self):
        """清空搜索"""
        self.search_var.set("")
        if self.current_detailed_data is not None:
            self.display_detailed_data(self.current_detailed_data)
        self.status_var.set("搜索已清空")

    def search_history_data(self):
        """搜索历史数据"""
        if self.history_purchase_data is None:
            messagebox.showwarning("警告", "请先加载历史采购数据")
            return

        keyword = self.history_search_var.get().strip()
        if not keyword:
            self.update_history_data_tab()
            return

        try:
            for item in self.history_tree.get_children():
                self.history_tree.delete(item)

            data = self.history_purchase_data.copy()

            mask = (
                    data['商品编码'].astype(str).str.contains(keyword, case=False, na=False) |
                    data['商品名称'].astype(str).str.contains(keyword, case=False, na=False)
            )

            filtered_data = data[mask]

            if len(filtered_data) == 0:
                self.history_info_label.config(text=f"未找到包含 '{keyword}' 的商品")
                return

            for _, row in filtered_data.iterrows():
                values = (
                    row.get('商品编码', ''),
                    row.get('商品名称', ''),
                    f"{row.get('采购单价', 0):,.2f}",
                    row.get('建单时间', '')
                )
                self.history_tree.insert('', tk.END, values=values)

            self.history_info_label.config(text=f"找到 {len(filtered_data)} 条匹配记录")

        except Exception as e:
            self.logger.error(f"搜索历史数据失败: {str(e)}")
            messagebox.showerror("搜索错误", f"搜索失败: {str(e)}")

    def clear_history_search(self):
        """清空历史数据搜索"""
        self.history_search_var.set("")
        self.update_history_data_tab()
        self.trend_info_label.config(text="请先搜索商品查看价格趋势")

        for widget in self.trend_canvas_frame.winfo_children():
            widget.destroy()

    def generate_price_trend(self):
        """生成价格趋势图"""
        keyword = self.history_search_var.get().strip()
        if not keyword:
            messagebox.showwarning("警告", "请输入商品编码或名称进行搜索")
            return

        if self.history_purchase_data is None:
            messagebox.showwarning("警告", "请先加载历史采购数据")
            return

        try:
            data = self.history_purchase_data.copy()

            mask = (
                    data['商品编码'].astype(str).str.contains(keyword, case=False, na=False) |
                    data['商品名称'].astype(str).str.contains(keyword, case=False, na=False)
            )

            product_data = data[mask]

            if len(product_data) == 0:
                messagebox.showwarning("警告", f"未找到商品 '{keyword}'")
                return

            product_code = product_data.iloc[0]['商品编码']
            product_name = product_data.iloc[0]['商品名称']

            product_history = data[data['商品编码'] == product_code].copy()

            if len(product_history) == 0:
                messagebox.showwarning("警告", f"商品 '{product_code}' 没有采购记录")
                return

            product_history = product_history.sort_values('建单时间')

            thirty_days_ago = datetime.now() - timedelta(days=30)
            recent_data = product_history[product_history['建单时间'] >= thirty_days_ago]

            if len(recent_data) == 0:
                recent_data = product_history

            for widget in self.trend_canvas_frame.winfo_children():
                widget.destroy()

            # 关键：设置中文字体
            plt.rcParams['font.sans-serif'] = self.font_fallbacks
            plt.rcParams['axes.unicode_minus'] = False

            fig = Figure(figsize=(8, 5), dpi=100)
            ax = fig.add_subplot(111)

            dates = recent_data['建单时间'].dt.strftime('%Y-%m-%d %H:%M').tolist()
            prices = recent_data['采购单价'].tolist()

            ax.plot(dates, prices, marker='o', linewidth=2, markersize=8, color='#2196F3')
            ax.fill_between(dates, prices, alpha=0.3, color='#2196F3')

            ax.set_title(f'{product_code} - {product_name} 价格趋势',
                         fontsize=14, fontweight='bold', pad=15)
            ax.set_xlabel('日期', fontsize=12)
            ax.set_ylabel('采购单价（元）', fontsize=12)
            ax.grid(True, alpha=0.3, linestyle='--')

            plt.setp(ax.xaxis.get_majorticklabels(), rotation=45, ha='right')

            for i, (date, price) in enumerate(zip(dates, prices)):
                ax.annotate(f'{price:.2f}',
                            (i, price),
                            textcoords="offset points",
                            xytext=(0, 10),
                            ha='center',
                            fontsize=9)

            fig.tight_layout()

            canvas = FigureCanvasTkAgg(fig, master=self.trend_canvas_frame)
            canvas.draw()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            toolbar = NavigationToolbar2Tk(canvas, self.trend_canvas_frame)
            toolbar.update()
            canvas.get_tk_widget().pack(fill=tk.BOTH, expand=True)

            self.trend_info_label.config(text=f"商品: {product_code} - {product_name} | 记录数: {len(recent_data)}")

            self.logger.info(f"生成了商品 {product_code} 的价格趋势图")

        except Exception as e:
            self.logger.error(f"生成价格趋势图失败: {str(e)}")
            messagebox.showerror("错误", f"生成价格趋势图失败: {str(e)}")

    # ================= 核心计算逻辑 =================
    def calculate_margin(self):
        """计算毛利"""
        sales_path = self.sales_file_path.get()
        latest_path = self.latest_purchase_file_path.get()

        if not sales_path or not latest_path:
            messagebox.showerror("错误", "请选择销售数据和最新采购数据文件")
            return

        if not self.history_loaded:
            pass

        try:
            self.update_progress(10, "开始数据处理...")
            self.logger.info("开始计算毛利...")

            self.update_progress(20, "读取销售数据...")

            sales_data = pd.read_excel(sales_path)
            sales_data = self.auto_rename_columns(sales_data, 'sales')

            if '销售日期' not in sales_data.columns:
                date_col_found = False
                for col in sales_data.columns:
                    if '日期' in str(col) or '时间' in str(col) or 'date' in str(col).lower():
                        try:
                            sales_data['销售日期'] = pd.to_datetime(sales_data[col], errors='coerce')
                            date_col_found = True
                            break
                        except:
                            pass
                if not date_col_found:
                    self.logger.warning("销售数据中未找到日期列，将使用当前系统日期。")
                    sales_data['销售日期'] = datetime.now()

            sales_data['销售日期'] = pd.to_datetime(sales_data['销售日期'], errors='coerce')

            latest_sales_date = sales_data['销售日期'].max()
            if pd.isna(latest_sales_date):
                latest_sales_date = datetime.now()
                self.logger.warning("无法从销售数据中获取有效日期，使用当前系统日期。")

            date_folder_str = latest_sales_date.strftime('%Y-%m')
            date_file_str = latest_sales_date.strftime('%Y-%m-%d')

            output_folder = os.path.join(self.config['export_path'], f"销售数据_{date_folder_str}")
            os.makedirs(output_folder, exist_ok=True)

            self.update_progress(40, "读取最新采购数据...")

            latest_data = pd.read_excel(latest_path)
            latest_data = self.auto_rename_columns(latest_data, 'purchase_latest')

            self.update_progress(45, "更新历史采购数据库...")
            latest_data['建单时间'] = datetime.now()

            if self.history_purchase_data is not None:
                all_purchase_data = pd.concat([self.history_purchase_data, latest_data], ignore_index=True)
            else:
                all_purchase_data = latest_data

            all_purchase_data = all_purchase_data.sort_values('建单时间', ascending=False)
            self.history_purchase_data = all_purchase_data.drop_duplicates('商品编码', keep='first')

            self.save_history_to_disk()

            record_count = len(self.history_purchase_data)
            unique_products = self.history_purchase_data['商品编码'].nunique()
            self.history_file_path.set(f"历史数据已更新: {record_count}条记录, {unique_products}个商品")
            self.update_history_data_tab()
            self.logger.info(f"历史采购数据已自动更新至: {record_count}条")

            self.update_progress(60, "合并采购价格...")

            latest_purchase_prices = self.history_purchase_data[['商品编码', '采购单价']]

            self.update_progress(70, "计算毛利...")

            merged_data = pd.merge(
                sales_data,
                latest_purchase_prices,
                on='商品编码',
                how='left'
            )

            merged_data['采购单价'] = merged_data['采购单价'].fillna(0)

            merged_data['销售金额'] = merged_data['订货数量'] * merged_data['商品单价']
            merged_data['采购成本'] = merged_data['订货数量'] * merged_data['采购单价']
            merged_data['销售毛利'] = merged_data['销售金额'] - merged_data['采购成本']
            merged_data['毛利率'] = np.where(
                merged_data['销售金额'] > 0,
                (merged_data['销售毛利'] / merged_data['销售金额']) * 100,
                0
            )

            numeric_cols = ['商品单价', '销售金额', '采购单价', '采购成本', '销售毛利']
            for col in numeric_cols:
                merged_data[col] = merged_data[col].round(2)
            merged_data['毛利率'] = merged_data['毛利率'].round(2)
            merged_data['订货数量'] = merged_data['订货数量'].astype(int)

            self.current_detailed_data = merged_data.copy()

            self.update_progress(85, "生成分析报告...")

            analysis_type = self.analysis_type.get()
            summaries = self.generate_comprehensive_analysis(merged_data, analysis_type, date_file_str)

            self.update_cumulative_data(merged_data, analysis_type, date_file_str)

            output_filename = f"销售毛利分析报告_{date_file_str}.xlsx"
            output_path = os.path.join(output_folder, output_filename)

            self.export_to_excel(merged_data, summaries, output_path, analysis_type, date_file_str)

            self.display_results(merged_data, summaries, output_path, date_file_str)

            if self.create_charts_var.get():
                self.generate_charts(merged_data, summaries)

            self.update_progress(95, "正在创建数据备份...")

            backup_path = os.path.join(output_folder, "原始数据备份")
            os.makedirs(backup_path, exist_ok=True)

            try:
                shutil.copy2(sales_path, os.path.join(backup_path, f"销售数据_{date_file_str}.xlsx"))
                shutil.copy2(latest_path, os.path.join(backup_path, f"最新采购数据_{date_file_str}.xlsx"))
            except Exception as e:
                self.logger.warning(f"备份原始数据失败: {e}")

            self.update_progress(100, "分析完成！")
            self.logger.info(f"分析完成，报告保存到: {output_path}")

            if self.config.get('auto_open', True):
                self.open_export_folder()

            messagebox.showinfo("成功", f"分析完成！\n历史采购数据已自动更新。\n文件已保存到:\n{output_path}")

        except Exception as e:
            self.logger.error(f"分析失败: {str(e)}", exc_info=True)
            messagebox.showerror("错误", f"分析失败: {str(e)}")

    def update_cumulative_data(self, data, analysis_type, date_str):
        """更新累计数据"""
        try:
            date = datetime.strptime(date_str, '%Y-%m-%d')
            month_key = date.strftime('%Y-%m')
            year_key = date.strftime('%Y')

            total_sales = data['销售金额'].sum()
            total_cost = data['采购成本'].sum()
            total_margin = data['销售毛利'].sum()
            margin_rate = (total_margin / total_sales * 100) if total_sales > 0 else 0
            product_count = data['商品编码'].nunique()

            if month_key not in self.monthly_data:
                self.monthly_data[month_key] = {
                    'total_sales': 0,
                    'total_cost': 0,
                    'total_margin': 0,
                    'margin_rate': 0,
                    'product_count': 0
                }

            self.monthly_data[month_key]['total_sales'] += total_sales
            self.monthly_data[month_key]['total_cost'] += total_cost
            self.monthly_data[month_key]['total_margin'] += total_margin

            if self.monthly_data[month_key]['total_sales'] > 0:
                self.monthly_data[month_key]['margin_rate'] = (
                        self.monthly_data[month_key]['total_margin'] /
                        self.monthly_data[month_key]['total_sales'] * 100
                )

            self.monthly_data[month_key]['product_count'] = max(
                self.monthly_data[month_key]['product_count'], product_count
            )

            if year_key not in self.yearly_data:
                self.yearly_data[year_key] = {
                    'total_sales': 0,
                    'total_cost': 0,
                    'total_margin': 0,
                    'margin_rate': 0,
                    'product_count': 0
                }

            self.yearly_data[year_key]['total_sales'] += total_sales
            self.yearly_data[year_key]['total_cost'] += total_cost
            self.yearly_data[year_key]['total_margin'] += total_margin

            if self.yearly_data[year_key]['total_sales'] > 0:
                self.yearly_data[year_key]['margin_rate'] = (
                        self.yearly_data[year_key]['total_margin'] /
                        self.yearly_data[year_key]['total_sales'] * 100
                )

            self.yearly_data[year_key]['product_count'] = max(
                self.yearly_data[year_key]['product_count'], product_count
            )

            self.save_cumulative_data()

            self.logger.info(f"累计数据已更新: 月度[{month_key}], 年度[{year_key}]")

        except Exception as e:
            self.logger.error(f"更新累计数据失败: {e}")

    def auto_rename_columns(self, df, data_type):
        """自动重命名列"""
        column_mapping = {
            'sales': {
                '商品编码': ['商品编码', '商品代码', '编码', '货号', 'SKU', 'Product Code'],
                '商品名称': ['商品名称', '商品名', '产品名称', '品名', 'Product Name'],
                '门店名称': ['门店名称', '店铺名称', '门店', '店铺', 'Store Name'],
                '一级分类': ['一级分类', '大类', '品类', '商品分类', 'Category'],
                '二级分类': ['二级分类', '小类', '子类', '商品子类', 'Sub Category'],
                '订货数量': ['订货数量', '数量', '销售数量', '销量', 'Qty', 'Quantity'],
                '商品单价': ['商品单价', '单价', '售价', '销售单价', 'Price', 'Unit Price'],
                '销售日期': ['销售日期', '日期', '业务日期', '下单日期', 'Date', 'Sale Date']
            },
            'purchase_latest': {
                '商品编码': ['商品编码', '商品代码', '编码', '货号', 'SKU', 'Product Code'],
                '采购单价': ['采购单价', '采购价', '进价', '成本价', 'Cost'],
                '商品名称': ['商品名称', '商品名', '产品名称', '品名', 'Product Name']
            },
            'purchase_history': {
                '商品编码': ['商品编码', '商品代码', '编码', '货号', 'SKU', 'Product Code'],
                '采购单价': ['采购单价', '采购价', '进价', '成本价', 'Cost'],
                '建单时间': ['建单时间', '创建时间', '下单时间', '时间', '日期', 'Date'],
                '商品名称': ['商品名称', '商品名', '产品名称', '品名', 'Product Name']
            }
        }

        mapping = column_mapping.get(data_type, {})
        df_columns_lower = {str(col).lower(): col for col in df.columns}

        for target_col, possible_names in mapping.items():
            if target_col not in df.columns:
                for name in possible_names:
                    if name.lower() in df_columns_lower:
                        original_col = df_columns_lower[name.lower()]
                        df = df.rename(columns={original_col: target_col})
                        break

        return df

    def generate_comprehensive_analysis(self, data, analysis_type, date_str):
        """生成全面的分析报告"""
        summaries = {}

        date = datetime.strptime(date_str, '%Y-%m-%d')
        data_date = date.strftime('%Y年%m月%d日')
        month_key = date.strftime('%Y-%m')
        year_key = date.strftime('%Y')

        total_sales = data['销售金额'].sum()
        total_cost = data['采购成本'].sum()
        total_margin = data['销售毛利'].sum()

        type_mapping = {'daily': '当日', 'monthly': '月度累计', 'yearly': '年度累计'}
        analysis_type_text = type_mapping[analysis_type]

        if analysis_type == 'monthly' and month_key in self.monthly_data:
            hist = self.monthly_data[month_key]
            total_sales = hist.get('total_sales', 0)
            total_cost = hist.get('total_cost', 0)
            total_margin = hist.get('total_margin', 0)
            data_date = f"{month_key}月度累计"
        elif analysis_type == 'yearly' and year_key in self.yearly_data:
            hist = self.yearly_data[year_key]
            total_sales = hist.get('total_sales', 0)
            total_cost = hist.get('total_cost', 0)
            total_margin = hist.get('total_margin', 0)
            data_date = f"{year_key}年度累计"

        total_summary = {
            '数据日期': data_date,
            '分析类型': f"{analysis_type_text}",
            '总销售金额（元）': total_sales,
            '总采购成本（元）': total_cost,
            '总销售毛利（元）': total_margin,
            '综合毛利率': (total_margin / total_sales * 100) if total_sales > 0 else 0,
            '商品种类数': data['商品编码'].nunique(),
            '门店数量': data['门店名称'].nunique() if '门店名称' in data.columns else 0,
            '总记录数': len(data),
            '平均毛利率': data['毛利率'].mean(),
            '毛利率中位数': data['毛利率'].median()
        }

        if analysis_type == 'monthly':
            total_summary['分析周期'] = '本月累计'
        elif analysis_type == 'yearly':
            total_summary['分析周期'] = '本年累计'
        else:
            total_summary['分析周期'] = '当日'

        summaries['总体情况'] = total_summary

        if '门店名称' in data.columns:
            store_summary = data.groupby('门店名称').agg({
                '销售金额': 'sum',
                '采购成本': 'sum',
                '销售毛利': 'sum',
                '商品编码': 'nunique',
                '订货数量': 'sum'
            }).reset_index()

            store_summary['毛利率'] = (store_summary['销售毛利'] / store_summary['销售金额'] * 100).round(2)
            store_summary = store_summary.sort_values('销售毛利', ascending=False)
            store_summary = store_summary.rename(columns={
                '商品编码': '商品种类',
                '订货数量': '销售数量',
                '销售金额': '销售金额（元）',
                '采购成本': '采购成本（元）',
                '销售毛利': '销售毛利（元）'
            })

            summaries['门店分析'] = store_summary

        if '一级分类' in data.columns:
            category_summary = data.groupby('一级分类').agg({
                '销售金额': 'sum',
                '采购成本': 'sum',
                '销售毛利': 'sum',
                '商品编码': 'nunique'
            }).reset_index()

            category_summary['毛利率'] = (category_summary['销售毛利'] / category_summary['销售金额'] * 100).round(2)
            category_summary = category_summary.sort_values('销售毛利', ascending=False)
            category_summary = category_summary.rename(columns={
                '商品编码': '商品种类',
                '销售金额': '销售金额（元）',
                '采购成本': '采购成本（元）',
                '销售毛利': '销售毛利（元）'
            })

            summaries['分类分析'] = category_summary

        margin_bins = [-float('inf'), 0, 10, 20, 30, 50, float('inf')]
        margin_labels = ['亏损', '0-10%', '10-20%', '20-30%', '30-50%', '50%以上']

        data['毛利率区间'] = pd.cut(data['毛利率'], bins=margin_bins, labels=margin_labels)
        margin_dist = data['毛利率区间'].value_counts().sort_index().reset_index()
        margin_dist.columns = ['毛利率区间', '商品数量']
        margin_dist['占比'] = (margin_dist['商品数量'] / len(data) * 100).round(2)

        summaries['毛利率分布'] = margin_dist

        top_products = data.nlargest(20, '销售毛利')[['商品编码', '商品名称', '销售金额', '销售毛利', '毛利率']].copy()
        top_products.index = range(1, len(top_products) + 1)
        top_products = top_products.rename(columns={
            '销售金额': '销售金额（元）',
            '销售毛利': '销售毛利（元）'
        })
        summaries['TOP商品'] = top_products

        if '门店名称' in data.columns:
            efficiency = data.groupby('门店名称').agg({
                '销售金额': 'sum',
                '销售毛利': 'sum',
                '商品编码': 'nunique'
            }).reset_index()

            efficiency['坪效'] = (efficiency['销售金额'] / efficiency['商品编码']).round(2)
            efficiency['毛利贡献率'] = (efficiency['销售毛利'] / total_margin * 100).round(2)
            efficiency = efficiency.sort_values('坪效', ascending=False)
            efficiency = efficiency.rename(columns={
                '销售金额': '销售金额（元）',
                '销售毛利': '销售毛利（元）',
                '商品编码': '商品种类'
            })

            summaries['效率分析'] = efficiency

        return summaries

    def export_to_excel(self, data, summaries, output_path, analysis_type, date_str):
        """导出到Excel"""

        book = Workbook()
        book.remove(book.active)

        with pd.ExcelWriter(output_path, engine='openpyxl', book=book) as writer:
            detailed_cols = ['商品编码', '商品名称', '门店名称', '一级分类', '二级分类',
                             '订货数量', '商品单价（元）', '销售金额（元）', '采购单价（元）', '采购成本（元）',
                             '销售毛利（元）', '毛利率']

            data_export = data.copy()
            rename_dict = {
                '商品单价': '商品单价（元）',
                '销售金额': '销售金额（元）',
                '采购单价': '采购单价（元）',
                '采购成本': '采购成本（元）',
                '销售毛利': '销售毛利（元）'
            }
            data_export = data_export.rename(columns=rename_dict)

            available_cols = [col for col in detailed_cols if col in data_export.columns]
            data_export[available_cols].to_excel(writer, sheet_name='详细数据', index=False)

            total_df = pd.DataFrame([summaries['总体情况']])
            total_df.to_excel(writer, sheet_name='总体情况', index=False)

            if '门店分析' in summaries:
                summaries['门店分析'].to_excel(writer, sheet_name='门店分析', index=False)

            if '分类分析' in summaries:
                summaries['分类分析'].to_excel(writer, sheet_name='分类分析', index=False)

            if '毛利率分布' in summaries:
                summaries['毛利率分布'].to_excel(writer, sheet_name='毛利率分布', index=False)

            if 'TOP商品' in summaries:
                summaries['TOP商品'].to_excel(writer, sheet_name='TOP商品', index=True)

            if '效率分析' in summaries:
                summaries['效率分析'].to_excel(writer, sheet_name='效率分析', index=False)

            if '销售日期' in data.columns:
                try:
                    daily_trend = data.groupby('销售日期').agg({
                        '销售金额': 'sum',
                        '采购成本': 'sum',
                        '销售毛利': 'sum'
                    }).reset_index()
                    daily_trend['毛利率'] = (daily_trend['销售毛利'] / daily_trend['销售金额'] * 100).round(2)
                    daily_trend = daily_trend.sort_values('销售日期')
                    daily_trend = daily_trend.rename(columns={
                        '销售金额': '销售金额（元）',
                        '采购成本': '采购成本（元）',
                        '销售毛利': '销售毛利（元）'
                    })
                    daily_trend.to_excel(writer, sheet_name='每日趋势', index=False)
                except Exception as e:
                    self.logger.error(f"生成每日趋势失败: {e}")

            cumulative_sheet_name = {
                'daily': '当日汇总',
                'monthly': '月度累计',
                'yearly': '年度累计'
            }[analysis_type]

            cumulative_data = {
                '项目': ['销售金额（元）', '采购成本（元）', '销售毛利（元）', '毛利率'],
                '金额': [
                    summaries['总体情况']['总销售金额（元）'],
                    summaries['总体情况']['总采购成本（元）'],
                    summaries['总体情况']['总销售毛利（元）'],
                    summaries['总体情况']['综合毛利率']
                ],
                '说明': ['所有商品销售总额', '所有商品采购成本', '销售利润总额', '综合利润率']
            }

            cumulative_df = pd.DataFrame(cumulative_data)
            cumulative_df.to_excel(writer, sheet_name=cumulative_sheet_name, index=False)

            if self.monthly_data:
                monthly_data_list = []
                sorted_months = sorted(self.monthly_data.keys())
                for month in sorted_months:
                    data_dict = self.monthly_data[month]
                    monthly_data_list.append({
                        '月份': month,
                        '销售金额（元）': data_dict['total_sales'],
                        '销售毛利（元）': data_dict['total_margin'],
                        '毛利率': data_dict['margin_rate'],
                        '商品种类数': data_dict['product_count']
                    })
                monthly_df = pd.DataFrame(monthly_data_list)
                monthly_df.to_excel(writer, sheet_name='月度累计数据', index=False)

            if self.yearly_data:
                yearly_data_list = []
                sorted_years = sorted(self.yearly_data.keys())
                for year in sorted_years:
                    data_dict = self.yearly_data[year]
                    yearly_data_list.append({
                        '年份': year,
                        '销售金额（元）': data_dict['total_sales'],
                        '销售毛利（元）': data_dict['total_margin'],
                        '毛利率': data_dict['margin_rate'],
                        '商品种类数': data_dict['product_count']
                    })
                yearly_df = pd.DataFrame(yearly_data_list)
                yearly_df.to_excel(writer, sheet_name='年度累计数据', index=False)

            source_info = {
                '项目': ['销售数据源', '最新采购数据源', '历史采购数据', '分析日期', '数据日期', '分析类型'],
                '内容': [
                    os.path.basename(self.sales_file_path.get()) if self.sales_file_path.get() else '未选择',
                    os.path.basename(
                        self.latest_purchase_file_path.get()) if self.latest_purchase_file_path.get() else '未选择',
                    f"{len(self.history_purchase_data)}条记录" if self.history_purchase_data is not None else '未加载',
                    datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
                    date_str,
                    {'daily': '当日分析', 'monthly': '月度累计', 'yearly': '年度累计'}[analysis_type]
                ]
            }

            source_df = pd.DataFrame(source_info)
            source_df.to_excel(writer, sheet_name='数据来源', index=False)

            self.apply_excel_styles(writer.book, summaries)

        return output_path

    def apply_excel_styles(self, workbook, summaries):
        """应用Excel样式"""
        # 使用 Excel 默认中文字体，确保跨平台兼容性 (通常 Calibri / Arial)
        # 如果必须用特定字体，可以使用 name='PingFang SC'，但在非 Mac Excel 上可能会显示为替代字体

        for sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]

            for column in worksheet.columns:
                column_letter = get_column_letter(column[0].column)
                first_cell_value = column[0].value

                if first_cell_value and '商品编码' in str(first_cell_value):
                    for cell in column:
                        cell.number_format = '@'

            for column in worksheet.columns:
                max_length = 0
                column_letter = get_column_letter(column[0].column)

                for cell in column:
                    try:
                        if cell.value:
                            cell_value = str(cell.value)
                            if isinstance(cell.value, (int, float)):
                                if '毛利率' in str(cell.column) or '占比' in str(cell.column) or '率' in str(
                                        cell.column):
                                    cell_value = f"{cell.value:.2f}%"
                                elif any(keyword in str(cell.column) for keyword in ['金额', '成本', '毛利', '单价']):
                                    cell_value = f"¥{cell.value:,.2f}"
                                else:
                                    cell_value = f"{cell.value:,.0f}"

                            cell_length = len(cell_value)
                            if cell_length > max_length:
                                max_length = cell_length
                    except:
                        pass

                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column_letter].width = adjusted_width

            if worksheet.max_row > 0:
                for cell in worksheet[1]:
                    cell.font = Font(bold=True, color="FFFFFF", size=11)
                    cell.fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")

            for row in worksheet.iter_rows(min_row=2):
                for cell in row:
                    if isinstance(cell.value, (int, float)):
                        if '毛利率' in str(cell.column) or '占比' in str(cell.column) or '率' in str(cell.column):
                            cell.number_format = '0.00"%"'

                            if cell.value < 0:
                                cell.font = Font(color="FF0000", bold=True)
                            elif cell.value > 0.2:
                                cell.font = Font(color="00B050", bold=True)
                        elif any(keyword in str(cell.column) for keyword in ['金额', '成本', '毛利', '单价']):
                            cell.number_format = '¥#,##0.00'
                            if cell.value > 10000:
                                cell.font = Font(bold=True)
                        else:
                            cell.number_format = '#,##0'

                    if isinstance(cell.value, (int, float)):
                        cell.alignment = Alignment(horizontal="right", vertical="center")
                    else:
                        cell.alignment = Alignment(horizontal="left", vertical="center")

            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            for row in worksheet.iter_rows():
                for cell in row:
                    cell.border = thin_border

            if worksheet.max_row > 1:
                worksheet.freeze_panes = 'A2'

            if sheet_name == '分类分析' and '分类分析' in summaries:
                last_row = worksheet.max_row
                if last_row > 1:
                    for col in range(1, worksheet.max_column + 1):
                        cell = worksheet.cell(row=1, column=col)
                        if cell.value and '毛利率' in str(cell.value):
                            col_letter = get_column_letter(col)
                            data_range = f"{col_letter}2:{col_letter}{last_row}"

                            rule = DataBarRule(start_type="num", start_value=0,
                                               end_type="num", end_value=50,
                                               color="FF6384")
                            worksheet.conditional_formatting.add(data_range, rule)
                            break

            worksheet.row_dimensions[1].height = 25

    # ================= 结果展示方法 =================
    def display_results(self, data, summaries, output_path, date_str):
        """显示结果"""
        self.display_detailed_data(data)
        self.display_summary_analysis(summaries, output_path, date_str)

    def display_detailed_data(self, data):
        """显示详细数据"""
        self.clear_treeview()

        display_data = data.head(200)

        for idx, row in display_data.iterrows():
            values = (
                idx + 1,
                str(row.get('商品编码', '')),
                row.get('商品名称', ''),
                row.get('门店名称', ''),
                row.get('一级分类', ''),
                row.get('二级分类', ''),
                f"{row.get('订货数量', 0):,}",
                f"{row.get('商品单价', 0):,.2f}",
                f"{row.get('销售金额', 0):,.2f}",
                f"{row.get('采购单价', 0):,.2f}",
                f"{row.get('采购成本', 0):,.2f}",
                f"{row.get('销售毛利', 0):,.2f}",
                f"{row.get('毛利率', 0):.2f}%"
            )
            self.tree.insert('', tk.END, values=values)

    def display_summary_analysis(self, summaries, output_path, date_str):
        """显示汇总分析"""
        self.summary_text.delete(1.0, tk.END)

        total = summaries['总体情况']

        self.summary_text.insert(tk.END, "销售毛利分析报告\n", 'title')
        self.summary_text.insert(tk.END, f"数据日期: {total['数据日期']}\n\n", 'data')

        self.summary_text.insert(tk.END, "📊 总体经营情况\n", 'section')
        self.summary_text.insert(tk.END, "-" * 50 + "\n")

        self.summary_text.insert(tk.END, f"分析类型: {total['分析类型']}\n", 'data')
        self.summary_text.insert(tk.END, f"总销售金额: ", 'data')
        self.summary_text.insert(tk.END, f"¥{total['总销售金额（元）']:,.2f}\n", 'highlight')

        self.summary_text.insert(tk.END, f"总采购成本: ", 'data')
        self.summary_text.insert(tk.END, f"¥{total['总采购成本（元）']:,.2f}\n", 'highlight')

        self.summary_text.insert(tk.END, f"总销售毛利: ", 'data')
        self.summary_text.insert(tk.END, f"¥{total['总销售毛利（元）']:,.2f}\n", 'highlight')

        self.summary_text.insert(tk.END, f"综合毛利率: ", 'data')
        self.summary_text.insert(tk.END, f"{total['综合毛利率']:.2f}%\n", 'highlight')

        self.summary_text.insert(tk.END, f"商品种类: {total['商品种类数']:,} 种\n", 'data')
        self.summary_text.insert(tk.END, f"门店数量: {total['门店数量']:,} 家\n", 'data')
        self.summary_text.insert(tk.END, f"数据记录: {total['总记录数']:,} 条\n\n", 'data')

        if '门店分析' in summaries:
            self.summary_text.insert(tk.END, "🏆 门店毛利排行榜\n", 'section')
            self.summary_text.insert(tk.END, "-" * 50 + "\n")

            for i, (_, row) in enumerate(summaries['门店分析'].head(5).iterrows(), 1):
                self.summary_text.insert(tk.END, f"{i}. {row['门店名称']}: ", 'data')
                self.summary_text.insert(tk.END, f"毛利¥{row['销售毛利（元）']:,.2f} ", 'highlight')
                self.summary_text.insert(tk.END, f"(毛利率{row['毛利率']:.1f}%)\n", 'data')
            self.summary_text.insert(tk.END, "\n")

        if '分类分析' in summaries:
            self.summary_text.insert(tk.END, "📦 品类毛利排行榜\n", 'section')
            self.summary_text.insert(tk.END, "-" * 50 + "\n")

            for i, (_, row) in enumerate(summaries['分类分析'].head(5).iterrows(), 1):
                self.summary_text.insert(tk.END, f"{i}. {row['一级分类']}: ", 'data')
                self.summary_text.insert(tk.END, f"毛利¥{row['销售毛利（元）']:,.2f} ", 'highlight')
                self.summary_text.insert(tk.END, f"(毛利率{row['毛利率']:.1f}%)\n", 'data')
            self.summary_text.insert(tk.END, "\n")

        self.summary_text.insert(tk.END, "📈 累计数据统计\n", 'section')
        self.summary_text.insert(tk.END, "-" * 50 + "\n")

        date = datetime.strptime(date_str, '%Y-%m-%d')
        month_key = date.strftime('%Y-%m')
        year_key = date.strftime('%Y')

        if month_key in self.monthly_data:
            monthly = self.monthly_data[month_key]
            self.summary_text.insert(tk.END, f"月度累计 ({month_key}):\n", 'data')
            self.summary_text.insert(tk.END, f"  销售金额: ¥{monthly['total_sales']:,.2f}\n", 'data')
            self.summary_text.insert(tk.END, f"  销售毛利: ¥{monthly['total_margin']:,.2f}\n", 'data')
            self.summary_text.insert(tk.END, f"  毛利率: {monthly['margin_rate']:.2f}%\n\n", 'data')

        if year_key in self.yearly_data:
            yearly = self.yearly_data[year_key]
            self.summary_text.insert(tk.END, f"年度累计 ({year_key}):\n", 'data')
            self.summary_text.insert(tk.END, f"  销售金额: ¥{yearly['total_sales']:,.2f}\n", 'data')
            self.summary_text.insert(tk.END, f"  销售毛利: ¥{yearly['total_margin']:,.2f}\n", 'data')
            self.summary_text.insert(tk.END, f"  毛利率: {yearly['margin_rate']:.2f}%\n\n", 'data')

        self.summary_text.insert(tk.END, "💾 报告信息\n", 'section')
        self.summary_text.insert(tk.END, "-" * 50 + "\n")
        self.summary_text.insert(tk.END, f"报告路径: {output_path}\n", 'data')
        self.summary_text.insert(tk.END, f"生成时间: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n", 'data')

    # ================= 图表生成方法 =================
    def generate_charts(self, data, summaries):
        """生成四宫格图表"""
        for frame in [self.chart_frame1, self.chart_frame2, self.chart_frame3, self.chart_frame4]:
            for widget in frame.winfo_children():
                if not isinstance(widget, tk.Label):
                    widget.destroy()

        # 关键：设置中文字体回退列表
        plt.rcParams['font.sans-serif'] = self.font_fallbacks
        plt.rcParams['axes.unicode_minus'] = False

        colors = ['#FF6B6B', '#4ECDC4', '#45B7D1', '#96CEB4', '#FFEAA7', '#DDA0DD', '#98D8C8', '#F7DC6F']

        try:
            if '一级分类' in data.columns:
                fig1 = self.create_sales_pie_chart(data, colors)
                canvas1 = FigureCanvasTkAgg(fig1, master=self.chart_frame1)
                canvas1.draw()
                canvas1.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            else:
                tk.Label(self.chart_frame1, text="无分类数据", font=(self.default_font, 12),
                         bg=self.colors['white']).pack(expand=True)

            if '一级分类' in data.columns:
                fig2 = self.create_margin_pie_chart(data, colors)
                canvas2 = FigureCanvasTkAgg(fig2, master=self.chart_frame2)
                canvas2.draw()
                canvas2.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            else:
                tk.Label(self.chart_frame2, text="无分类数据", font=(self.default_font, 12),
                         bg=self.colors['white']).pack(expand=True)

            if '门店分析' in summaries:
                fig3 = self.create_store_sales_bar_chart(summaries['门店分析'], colors)
                canvas3 = FigureCanvasTkAgg(fig3, master=self.chart_frame3)
                canvas3.draw()
                canvas3.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            else:
                tk.Label(self.chart_frame3, text="无门店数据", font=(self.default_font, 12),
                         bg=self.colors['white']).pack(expand=True)

            if '分类分析' in summaries:
                fig4 = self.create_category_margin_bar_chart(summaries['分类分析'], colors)
                canvas4 = FigureCanvasTkAgg(fig4, master=self.chart_frame4)
                canvas4.draw()
                canvas4.get_tk_widget().pack(fill=tk.BOTH, expand=True)
            else:
                tk.Label(self.chart_frame4, text="无分类数据", font=(self.default_font, 12),
                         bg=self.colors['white']).pack(expand=True)

            self.notebook.select(1)

            self.logger.info("四宫格图表生成完成")

        except Exception as e:
            self.logger.error(f"生成图表失败: {str(e)}")
            for frame in [self.chart_frame1, self.chart_frame2, self.chart_frame3, self.chart_frame4]:
                for widget in frame.winfo_children():
                    if not isinstance(widget, tk.Label):
                        widget.destroy()
                tk.Label(frame, text=f"图表生成失败\n{str(e)}", font=(self.default_font, 10),
                         bg=self.colors['white'], fg=self.colors['danger']).pack(expand=True)

    def create_sales_pie_chart(self, data, colors):
        """创建一级分类销售占比圆饼图"""
        category_sales = data.groupby('一级分类')['销售金额'].sum()

        total_sales = category_sales.sum()
        threshold = total_sales * 0.01
        main_categories = category_sales[category_sales >= threshold]
        other_sales = category_sales[category_sales < threshold].sum()

        if other_sales > 0:
            main_categories = pd.concat([main_categories, pd.Series({'其他': other_sales})])

        fig, ax = plt.subplots(figsize=(5, 4), dpi=100)

        wedges, texts, autotexts = ax.pie(main_categories.values,
                                          labels=main_categories.index,
                                          autopct='%1.1f%%',
                                          colors=colors[:len(main_categories)],
                                          startangle=90,
                                          pctdistance=0.85,
                                          textprops={'fontsize': 9})

        for autotext in autotexts:
            autotext.set_color('black')
            autotext.set_fontsize(8)
            autotext.set_fontweight('bold')

        ax.set_title('销售金额占比', fontsize=12, fontweight='bold', pad=15)

        ax.legend(wedges, main_categories.index,
                  title="分类",
                  loc="center left",
                  bbox_to_anchor=(1, 0, 0.5, 1),
                  fontsize=8)

        center_text = f"总销售额\n¥{total_sales:,.0f}"
        ax.text(0, 0, center_text, ha='center', va='center',
                fontsize=10, fontweight='bold', color='#333333')

        ax.axis('equal')

        plt.tight_layout()
        return fig

    def create_margin_pie_chart(self, data, colors):
        """创建一级分类毛利占比圆饼图"""
        category_margin = data.groupby('一级分类')['销售毛利'].sum()

        total_margin = category_margin.sum()
        threshold = abs(total_margin) * 0.01
        main_categories = category_margin[abs(category_margin) >= threshold]
        other_margin = category_margin[abs(category_margin) < threshold].sum()

        if abs(other_margin) > 0:
            main_categories = pd.concat([main_categories, pd.Series({'其他': other_margin})])

        fig, ax = plt.subplots(figsize=(5, 4), dpi=100)

        if len(main_categories) > 0:
            wedges, texts, autotexts = ax.pie(main_categories.values,
                                              labels=main_categories.index,
                                              autopct=lambda pct: f'{pct:.1f}%\n(¥{pct * total_margin / 100:,.0f})',
                                              colors=colors[:len(main_categories)],
                                              startangle=90,
                                              pctdistance=0.85,
                                              textprops={'fontsize': 8})

            for autotext in autotexts:
                autotext.set_color('black')
                autotext.set_fontsize(7)
                autotext.set_fontweight('bold')

            ax.set_title('销售毛利占比', fontsize=12, fontweight='bold', pad=15)

            ax.legend(wedges, main_categories.index,
                      title="分类",
                      loc="center left",
                      bbox_to_anchor=(1, 0, 0.5, 1),
                      fontsize=8)

            center_text = f"总毛利\n¥{total_margin:,.0f}"
            ax.text(0, 0, center_text, ha='center', va='center',
                    fontsize=10, fontweight='bold',
                    color='green' if total_margin >= 0 else 'red')

        ax.axis('equal')

        plt.tight_layout()
        return fig

    def create_store_sales_bar_chart(self, store_data, colors):
        """创建门店销售排行条形图"""
        fig, ax = plt.subplots(figsize=(5, 4), dpi=100)

        top_stores = store_data.head(8).copy()
        top_stores = top_stores.sort_values('销售金额（元）')

        bars = ax.barh(range(len(top_stores)), top_stores['销售金额（元）'],
                       color=colors[:len(top_stores)], alpha=0.8)

        ax.set_yticks(range(len(top_stores)))
        ax.set_yticklabels(top_stores['门店名称'], fontsize=9)

        ax.set_xlabel('销售金额（元）', fontsize=10)

        ax.set_title('门店销售金额排行', fontsize=12, fontweight='bold', pad=15)

        ax.grid(True, alpha=0.3, axis='x', linestyle='--')

        for i, (bar, value) in enumerate(zip(bars, top_stores['销售金额（元）'])):
            width = bar.get_width()
            ax.text(width + width * 0.01, bar.get_y() + bar.get_height() / 2,
                    f'¥{value:,.0f}', va='center', fontsize=8, fontweight='bold')

        plt.tight_layout()

        return fig

    def create_category_margin_bar_chart(self, category_data, colors):
        """创建品类毛利排行条形图"""
        fig, ax = plt.subplots(figsize=(5, 4), dpi=100)

        top_categories = category_data.head(8).copy()
        top_categories = top_categories.sort_values('销售毛利（元）')

        bars = ax.barh(range(len(top_categories)), top_categories['销售毛利（元）'],
                       color=colors[:len(top_categories)], alpha=0.8)

        ax.set_yticks(range(len(top_categories)))
        ax.set_yticklabels(top_categories['一级分类'], fontsize=9)

        ax.set_xlabel('销售毛利（元）', fontsize=10)

        ax.set_title('品类销售毛利排行', fontsize=12, fontweight='bold', pad=15)

        ax.grid(True, alpha=0.3, axis='x', linestyle='--')

        for i, (bar, value) in enumerate(zip(bars, top_categories['销售毛利（元）'])):
            width = bar.get_width()
            ax.text(width + width * 0.01, bar.get_y() + bar.get_height() / 2,
                    f'¥{value:,.0f}', va='center', fontsize=8, fontweight='bold')

            margin_rate = top_categories.iloc[i]['毛利率']
            ax.text(bar.get_x() - bar.get_width() * 0.05, bar.get_y() + bar.get_height() / 2,
                    f'{margin_rate:.1f}%', va='center', fontsize=7, fontweight='bold',
                    color='green' if margin_rate >= 0 else 'red')

        plt.tight_layout()

        return fig

    def clear_treeview(self):
        """清空Treeview"""
        for item in self.tree.get_children():
            self.tree.delete(item)

    def update_progress(self, value, message):
        """更新进度"""
        self.progress_var.set(value)
        self.status_label.config(text=message)
        self.status_var.set(message)
        self.root.update()


def main():
    root = tk.Tk()

    root.update_idletasks()
    width = 1400
    height = 900
    screen_width = root.winfo_screenwidth()
    screen_height = root.winfo_screenheight()
    x = (screen_width - width) // 2
    y = (screen_height - height) // 2
    root.geometry(f'{width}x{height}+{x}+{y}')

    app = ProfessionalMarginCalculator(root)

    root.mainloop()


if __name__ == "__main__":
    main()